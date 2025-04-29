import unittest
from openpyxl import load_workbook
from openpyxl import Workbook

class TestExcel(unittest.TestCase):

    def test_gerar_arquivo_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Evento", "Data"])
        wb.save('test_eventos.xlsx')

        wb_carregado = load_workbook('test_eventos.xlsx')
        sheet = wb_carregado.active
        self.assertEqual(sheet['A1'].value, 'Evento')

if __name__ == '__main__':
    unittest.main()
